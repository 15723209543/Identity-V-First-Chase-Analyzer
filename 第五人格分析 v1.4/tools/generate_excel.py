# -*- coding: gb2312 -*-
from pathlib import Path
import re

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


root = Path(__file__).resolve().parents[1]


def read_text(path: Path) -> str:
    for encoding in ("utf-8", "gbk", "gb2312"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    return path.read_text(encoding="utf-8", errors="ignore")


def parse_tags(raw: str) -> str:
    return "、".join(re.findall(r'"([^"]+)"', raw))


def parse_survivors():
    text = read_text(root / "include" / "survivor_data.h")
    pattern = re.compile(
        r'\{(\d+), "([^"]+)", (\d+), (\d+), (\d+), (\d+), (\d+), \{([^}]*)\}, "([^"]*)"\}',
        re.S,
    )
    rows = []
    for match in pattern.finditer(text):
        rows.append(
            [
                int(match.group(1)),
                match.group(2),
                int(match.group(3)),
                int(match.group(4)),
                int(match.group(5)),
                int(match.group(6)),
                int(match.group(7)),
                parse_tags(match.group(8)),
                match.group(9),
            ]
        )
    return rows


def parse_hunters():
    text = read_text(root / "include" / "hunter_data.h")
    pattern = re.compile(
        r'\{(\d+), "([^"]+)", (\d+), (\d+), (\d+), (\d+), (\d+), \{([^}]*)\}, "([^"]*)"\}',
        re.S,
    )
    rows = []
    for match in pattern.finditer(text):
        rows.append(
            [
                int(match.group(1)),
                match.group(2),
                int(match.group(3)),
                int(match.group(4)),
                int(match.group(5)),
                int(match.group(6)),
                int(match.group(7)),
                parse_tags(match.group(8)),
                match.group(9),
            ]
        )
    return rows


def parse_maps():
    text = read_text(root / "include" / "map_data.h")
    rows = []
    current_map_id = None
    current_map_name = None
    for line in text.splitlines():
        map_match = re.match(r'\s*\{(\d+), "([^"]+)", \d+, \d+, \{', line)
        if map_match:
            current_map_id = int(map_match.group(1))
            current_map_name = map_match.group(2)
            continue

        region_match = re.match(
            r'\s*\{(\d+), "([^"]+)", (\d+), (\d+), (\d+), (\d+), (\d+), (\d+), (\d+), "([^"]*)"\}',
            line,
        )
        if region_match and current_map_id is not None:
            rows.append(
                [
                    current_map_id,
                    current_map_name,
                    int(region_match.group(1)),
                    region_match.group(2),
                    int(region_match.group(3)),
                    int(region_match.group(4)),
                    int(region_match.group(5)),
                    int(region_match.group(6)),
                    int(region_match.group(7)),
                    int(region_match.group(8)),
                    int(region_match.group(9)),
                    region_match.group(10),
                ]
            )
    return rows


def build_map_stats(map_rows):
    stats = {}
    for row in map_rows:
        map_id, map_name = row[0], row[1]
        stats.setdefault((map_id, map_name), 0)
        stats[(map_id, map_name)] += 1

    rows = []
    for (map_id, map_name), count in sorted(stats.items()):
        rows.append(
            [
                map_id,
                map_name,
                "大图" if count == 12 else "小图",
                count,
                "排位选点方位",
                "符合要求" if count in (9, 12) else "需要检查",
            ]
        )
    return rows


def add_sheet(wb, title, headers, rows):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for column_cells in ws.columns:
        max_len = 0
        column = column_cells[0].column
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[get_column_letter(column)].width = min(max(max_len + 2, 10), 45)

    ws.freeze_panes = "A2"
    return ws


def main():
    wb = Workbook()
    wb.remove(wb.active)

    add_sheet(
        wb,
        "求生者",
        ["编号", "名称", "破译", "牵制", "辅助", "救援", "首抓价值", "标签", "建模说明"],
        parse_survivors(),
    )
    add_sheet(
        wb,
        "监管者",
        ["编号", "名称", "追击", "守椅", "控场", "信息", "综合强度", "标签", "建模说明"],
        parse_hunters(),
    )
    map_rows = parse_maps()
    add_sheet(
        wb,
        "地图区域",
        ["地图编号", "地图", "区域编号", "区域", "X", "Y", "宽", "高", "牵制分", "开阔分", "电机价值", "区域说明"],
        map_rows,
    )
    add_sheet(
        wb,
        "地图统计",
        ["地图编号", "地图", "类型", "区域数量", "方位口径", "检查结果"],
        build_map_stats(map_rows),
    )
    add_sheet(
        wb,
        "模型权重",
        ["项目", "权重", "说明"],
        [
            ["角色指标分", 1.00, "破译、牵制、辅助、救援、首抓价值折算"],
            ["监管者风格分", 0.80, "追击、守椅、控场、信息与目标角色类型匹配"],
            ["地图区域分", 0.55, "低牵制区、开阔区、电机价值高会提高意愿"],
            ["距离分", 0.70, "跨图距离越远，首追意愿越低"],
            ["克制分", 1.35, "严重角色克制权重最高"],
            ["概率转换", "softmax/18", "将 4 个 raw_score 转换为概率"],
        ],
    )
    add_sheet(
        wb,
        "输出设置",
        ["项目", "值", "说明"],
        [
            ["结果目录", "result", "程序每次输出时会自动创建该文件夹"],
            ["文件名格式", "YYYY-MM-DD_HH-MM-SS.txt", "用输出时的本地时间命名，避免 Windows 文件名中的冒号"],
            ["覆盖策略", "不覆盖", "每次运行都会生成新的时间文件"],
            ["右侧显示", "开启", "完整分析同时显示在 EasyX 右侧结果区"],
            ["输入方式", "鼠标点击", "地图、区域、求生者、监管者都在 EasyX 窗口中点击选择"],
            ["地图方位", "排位选点方位", "左上、右上、左下、右下对应排位区域选择图的大方位"],
        ],
    )
    add_sheet(
        wb,
        "克制规则",
        ["规则", "分值", "说明"],
        [
            ["anti_mobility vs mobility", 13, "反位移监管者更容易处理位移型求生者"],
            ["anti_item vs item_dependent", 15, "反道具能力压低道具牵制收益"],
            ["anti_heal vs healer", 16, "反治疗能力压低治疗体系收益"],
            ["anti_decoder vs decoder", 15, "反破译能力提高追高破译角色收益"],
            ["摄影师 vs decoder", 18, "相中世界对高破译角色压制明显"],
            ["牙医 vs airborne", 18, "牙医对飞行/空中转点角色给强克制加分"],
            ["强牵制且监管者追击低", -9, "慢速监管者追强牵制位会降低首追意愿"],
        ],
    )
    add_sheet(
        wb,
        "资料来源",
        ["来源", "链接", "用途"],
        [
            ["维基百科：第五人格", "https://zh.wikipedia.org/wiki/第五人格_(游戏)", "角色列表口径"],
            ["第五人格 BWIKI：区域选择", "https://wiki.biligame.com/dwrg/区域选择", "9-12 区域选择机制与排位选点口径"],
            ["第五人格 BWIKI：地图", "https://wiki.biligame.com/dwrg/地图", "排位地图与地图名称口径"],
            ["第五人格 BWIKI：求生者", "https://wiki.biligame.com/dwrg/求生者", "求生者角色口径"],
            ["第五人格 BWIKI：监管者", "https://wiki.biligame.com/dwrg/监管者", "监管者角色口径"],
        ],
    )

    output = root / "data" / "identityv_scores.xlsx"
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    print(output)


if __name__ == "__main__":
    main()
