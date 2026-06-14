# -*- coding: gb2312 -*-
from pathlib import Path
import importlib.util
import re

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


root = Path(__file__).resolve().parents[1]


def read_text(path: Path) -> str:
    for encoding in ("utf-8", "gbk", "gb2312"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    return path.read_text(encoding="utf-8", errors="ignore")


def parse_tags(raw: str) -> str:
    return "、".join(re.findall(r'"([^"]+)"', raw))


def parse_survivors():
    text = read_text(root / "include" / "survivor_data.h")
    pattern = re.compile(
        r'\{(\d+), "([^"]+)", (\d+), (\d+), (\d+), (\d+), (\d+), \{([^}]*)\}, "([^"]*)"\}',
        re.S,
    )
    rows = []
    for match in pattern.finditer(text):
        rows.append(
            [
                int(match.group(1)),
                match.group(2),
                int(match.group(3)),
                int(match.group(4)),
                int(match.group(5)),
                int(match.group(6)),
                int(match.group(7)),
                parse_tags(match.group(8)),
                match.group(9),
            ]
        )
    return rows


def parse_hunters():
    text = read_text(root / "include" / "hunter_data.h")
    pattern = re.compile(
        r'\{(\d+), "([^"]+)", (\d+), (\d+), (\d+), (\d+), (\d+), \{([^}]*)\}, "([^"]*)"\}',
        re.S,
    )
    rows = []
    for match in pattern.finditer(text):
        rows.append(
            [
                int(match.group(1)),
                match.group(2),
                int(match.group(3)),
                int(match.group(4)),
                int(match.group(5)),
                int(match.group(6)),
                int(match.group(7)),
                parse_tags(match.group(8)),
                match.group(9),
            ]
        )
    return rows


def parse_maps():
    text = read_text(root / "include" / "map_data.h")
    rows = []
    current_map_id = None
    current_map_name = None
    for line in text.splitlines():
        map_match = re.match(r'\s*\{(\d+), "([^"]+)", \d+, \d+, \{', line)
        if map_match:
            current_map_id = int(map_match.group(1))
            current_map_name = map_match.group(2)
            continue

        region_match = re.match(
            r'\s*\{(\d+), "([^"]+)", (\d+), (\d+), (\d+), (\d+), (\d+), (\d+), (\d+), "([^"]*)"\}',
            line,
        )
        if region_match and current_map_id is not None:
            rows.append(
                [
                    current_map_id,
                    current_map_name,
                    int(region_match.group(1)),
                    region_match.group(2),
                    int(region_match.group(3)),
                    int(region_match.group(4)),
                    int(region_match.group(5)),
                    int(region_match.group(6)),
                    int(region_match.group(7)),
                    int(region_match.group(8)),
                    int(region_match.group(9)),
                    region_match.group(10),
                ]
            )
    return rows


def build_map_stats(map_rows):
    stats = {}
    for row in map_rows:
        map_id, map_name = row[0], row[1]
        stats.setdefault((map_id, map_name), 0)
        stats[(map_id, map_name)] += 1

    rows = []
    for (map_id, map_name), count in sorted(stats.items()):
        rows.append(
            [
                map_id,
                map_name,
                "大图" if count == 12 else "小图",
                count,
                "用户提供 data/map 排位选点方位",
                "符合要求" if count in (9, 12) else "需要检查",
            ]
        )
    return rows


def build_geometry_rows():
    script_path = root / "tools" / "generate_map_geometry.py"
    spec = importlib.util.spec_from_file_location("generate_map_geometry", script_path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)

    rows = []
    for item in module.maps:
        geometry = module.extract_map(item)
        rows.append(
            [
                geometry["id"],
                item["name"],
                geometry["source"],
                module.grid_width,
                module.grid_height,
                round(geometry["real_width"], 2),
                round(geometry["real_height"], 2),
                len(geometry["wall_cells"]),
                len(geometry["windows"]),
                len(geometry["pallets"]),
                "左侧直接显示用户原图；几何数据只用于寻路距离、白色空地复核和板窗资源评分",
            ]
        )
    return rows


def build_interface_rows():
    return [
        ["左侧地图", "原图显示", "EasyX 直接加载 data/map 中的 jpg，不再额外重画墙体、窗、板或大门红块"],
        ["有效地图范围", "按图源裁剪框限制", "只能在实际对局地图外墙附近的有效裁剪范围内点选，标题、图例、图片白边不参与点选"],
        ["可点击区域", "外墙内部白色空地", "点位必须通过原图白色像素、非边缘白底、几何网格非墙体三层判定"],
        ["不可点击区域", "有颜色或图外", "墙体、建筑、窗、板、地下室、门图标、文字、图例、外墙外白底都不可选择"],
        ["大门显示", "以原图为准", "不再额外叠加两个红色小块，门的位置直接看用户提供的地图原图"],
        ["角色选择", "右侧列表点击", "位置在左侧地图点击，求生者和监管者角色只能在右侧编号列表点击"],
        ["结果滚动", "鼠标滚轮", "完整分析在右侧显示，内容过长时用鼠标滚轮查看"],
        ["求生者输出模板", "三行", "编号+名称+出生地；能力参数；现在应该怎么做"],
        ["监管者输出模板", "固定结构", "编号+名称；能力参数+强度分析；每个求生者三行：名称+概率、计算系数原理、纯文字说明；第三行不重复第二行数据"],
    ]


def add_sheet(wb, title, headers, rows):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for column_cells in ws.columns:
        max_len = 0
        column = column_cells[0].column
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[get_column_letter(column)].width = min(max(max_len + 2, 10), 45)

    ws.freeze_panes = "A2"
    return ws


def main():
    wb = Workbook()
    wb.remove(wb.active)

    add_sheet(
        wb,
        "求生者",
        ["编号", "名称", "破译", "牵制", "辅助", "救援", "首抓价值", "标签", "建模说明"],
        parse_survivors(),
    )
    add_sheet(
        wb,
        "监管者",
        ["编号", "名称", "追击", "守椅", "控场", "信息", "综合强度", "标签", "建模说明"],
        parse_hunters(),
    )
    map_rows = parse_maps()
    add_sheet(
        wb,
        "地图区域",
        ["地图编号", "地图", "区域编号", "区域", "X", "Y", "宽", "高", "牵制分", "开阔分", "电机价值", "区域说明"],
        map_rows,
    )
    add_sheet(
        wb,
        "地图统计",
        ["地图编号", "地图", "类型", "区域数量", "方位口径", "检查结果"],
        build_map_stats(map_rows),
    )
    add_sheet(
        wb,
        "几何地图",
        ["地图编号", "地图键", "图源文件", "网格宽", "网格高", "寻路宽度m", "寻路高度m", "墙格数", "窗数量", "板数量", "说明"],
        build_geometry_rows(),
    )
    add_sheet(
        wb,
        "模型权重",
        ["项目", "权重", "说明"],
        [
            ["角色指标分", 1.00, "破译、牵制、辅助、救援、首抓价值折算"],
            ["监管者风格分", 0.80, "追击、守椅、控场、信息与目标角色类型匹配"],
            ["地图区域分", 0.55, "低牵制区、开阔区、电机价值高会提高意愿"],
            ["距离分", 0.70, "跨图距离越远，首追意愿越低"],
            ["克制分", 1.35, "严重角色克制权重最高"],
            ["概率转换", "softmax/18", "将 4 个 raw_score 转换为概率"],
        ],
    )
    add_sheet(
        wb,
        "输出设置",
        ["项目", "值", "说明"],
        [
            ["工程名", "D5_firstcaught", "Visual Studio 解决方案和工程名"],
            ["结果目录", "result", "程序每次输出时会自动创建该文件夹"],
            ["文件名格式", "YYYY-MM-DD_HH-MM-SS.txt", "用输出时的本地时间命名，避免 Windows 文件名中的冒号"],
            ["覆盖策略", "不覆盖", "每次运行都会生成新的时间文件"],
            ["右侧显示", "开启", "完整分析同时显示在 EasyX 右侧结果区"],
            ["输入方式", "鼠标点击", "地图位置只能点左侧地图白色空地，角色只能点右侧角色列表"],
            ["地图方位", "data/map 用户方位图", "左上、右上、左下、右下对应用户提供排位选点图的大方位"],
            ["几何寻路", "开启", "根据墙体网格搜索路径距离，不再使用区域中心直线距离"],
            ["点击复核", "开启", "点击点必须位于有效裁剪框、外墙内部白色空地，并且不能落在几何墙体上"],
            ["板窗加成", "开启", "点击点附近窗和板会降低该位置的首追风险"],
            ["大门标记", "原图显示", "不再额外绘制红色小块，直接使用地图原图上的门图标"],
            ["撤销按钮", "开启", "位置页可返回/撤销上一名，角色页可撤销当前点位"],
        ],
    )
    add_sheet(
        wb,
        "界面规则",
        ["项目", "当前规则", "说明"],
        build_interface_rows(),
    )
    add_sheet(
        wb,
        "克制规则",
        ["规则", "分值", "说明"],
        [
            ["anti_mobility vs mobility", 13, "反位移监管者更容易处理位移型求生者"],
            ["anti_item vs item_dependent", 15, "反道具能力压低道具牵制收益"],
            ["anti_heal vs healer", 16, "反治疗能力压低治疗体系收益"],
            ["anti_decoder vs decoder", 15, "反破译能力提高追高破译角色收益"],
            ["摄影师 vs decoder", 18, "相中世界对高破译角色压制明显"],
            ["牙医 vs airborne", 18, "牙医对飞行/空中转点角色给强克制加分"],
            ["强牵制且监管者追击低", -9, "慢速监管者追强牵制位会降低首追意愿"],
        ],
    )
    add_sheet(
        wb,
        "资料来源",
        ["来源", "链接", "用途"],
        [
            ["用户提供地图原图", "data/map/*.jpg", "9 张排位地图的左侧显示图源，同时用于白色空地点选复核"],
            ["tools/generate_map_geometry.py", "本项目脚本", "从用户地图原图中提取墙体网格、窗和板，用于寻路和板窗资源评分"],
            ["维基百科：第五人格", "https://zh.wikipedia.org/wiki/第五人格_(游戏)", "角色列表口径"],
            ["第五人格 BWIKI：区域选择", "https://wiki.biligame.com/dwrg/区域选择", "9-12 区域选择机制与排位选点口径"],
            ["第五人格 BWIKI：地图", "https://wiki.biligame.com/dwrg/地图", "排位地图与地图名称口径"],
            ["第五人格 BWIKI：求生者", "https://wiki.biligame.com/dwrg/求生者", "求生者角色口径"],
            ["第五人格 BWIKI：监管者", "https://wiki.biligame.com/dwrg/监管者", "监管者角色口径"],
        ],
    )

    output = root / "data" / "identityv_scores.xlsx"
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    print(output)


if __name__ == "__main__":
    main()
