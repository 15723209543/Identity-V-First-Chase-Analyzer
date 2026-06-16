# -*- coding: utf-8 -*-
from pathlib import Path
import importlib.util
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


root = Path(__file__).resolve().parents[1]


def read_text(path: Path) -> str:
    for encoding in ("utf-8", "gbk", "gb2312"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    return path.read_text(encoding="utf-8", errors="ignore")


def parse_tags(raw: str) -> str:
    return "、".join(re.findall(r'"([^"]+)"', raw))


def to_percent(raw_score: int) -> float:
    return round(60 + (raw_score - 1) * 40 / 9, 2)


def parse_role_file(file_name: str):
    text = read_text(root / "include" / file_name)
    pattern = re.compile(
        r'\{(\d+), "([^"]+)", (\d+), (\d+), (\d+), (\d+), (\d+), \{([^}]*)\}, "([^"]*)"\}',
        re.S,
    )
    rows = []
    for match in pattern.finditer(text):
        values = [int(match.group(i)) for i in range(3, 8)]
        row = [int(match.group(1)), match.group(2)]
        for value in values:
            row.extend([value, to_percent(value)])
        row.extend([parse_tags(match.group(8)), match.group(9)])
        rows.append(row)
    return rows


def parse_auxiliary_traits():
    text = read_text(root / "include" / "auxiliary_trait_data.h")
    pattern = re.compile(
        r'\{\s*(\d+), "([^"]+)", ([\-\d.]+), ([\-\d.]+), ([\-\d.]+), \{([^}]*)\},\s*"([^"]*)"\s*\}',
        re.S,
    )
    rows = []
    for match in pattern.finditer(text):
        rows.append([
            int(match.group(1)),
            match.group(2),
            float(match.group(3)),
            float(match.group(4)),
            float(match.group(5)),
            parse_tags(match.group(6)),
            match.group(7),
            "开局 CD 未到时不修正追击时间；CD 到后按技能标签修正首刀或击倒时间",
        ])
    return rows


def parse_maps():
    text = read_text(root / "include" / "map_data.h")
    rows = []
    current_map_id = None
    current_map_name = None
    for line in text.splitlines():
        map_match = re.match(r'\s*\{(\d+), "([^"]+)", \d+, \d+, \{', line)
        if map_match:
            current_map_id = int(map_match.group(1))
            current_map_name = map_match.group(2)
            continue

        region_match = re.match(
            r'\s*\{(\d+), "([^"]+)", (\d+), (\d+), (\d+), (\d+), (\d+), (\d+), (\d+), "([^"]*)"\}',
            line,
        )
        if region_match and current_map_id is not None:
            rows.append([
                current_map_id,
                current_map_name,
                int(region_match.group(1)),
                region_match.group(2),
                int(region_match.group(3)),
                int(region_match.group(4)),
                int(region_match.group(5)),
                int(region_match.group(6)),
                int(region_match.group(7)),
                int(region_match.group(8)),
                int(region_match.group(9)),
                region_match.group(10),
            ])
    return rows


def build_map_stats(map_rows):
    stats = {}
    for row in map_rows:
        stats.setdefault((row[0], row[1]), 0)
        stats[(row[0], row[1])] += 1
    return [[mid, name, "大图" if count == 12 else "小图", count, "按 data/map 当前图源方位识别"]
            for (mid, name), count in sorted(stats.items())]


def build_geometry_rows():
    script_path = root / "tools" / "image_recognition" / "generate_map_geometry.py"
    spec = importlib.util.spec_from_file_location("generate_map_geometry", script_path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)

    rows = []
    for item in module.maps:
        geometry = module.extract_map(item)
        rows.append([
            geometry["id"],
            item["name"],
            geometry["source"],
            module.grid_width,
            module.grid_height,
            round(geometry["real_width"], 2),
            round(geometry["real_height"], 2),
            len(geometry["wall_cells"]),
            len(geometry["windows"]),
            len(geometry["pallets"]),
            "用于白色空地点选复核、寻路距离和板窗资源评分",
        ])
    return rows


def parse_model_params():
    rows = []
    path = root / "config" / "model_params.txt"
    if not path.exists():
        return rows
    comment = ""
    for line in path.read_text(encoding="utf-8").splitlines():
        stripped = line.strip()
        if not stripped:
            continue
        if stripped.startswith("#"):
            content = stripped.lstrip("#").strip()
            if content and not set(content) <= {"="}:
                comment = content
            continue
        if "=" not in stripped:
            continue
        key, value = stripped.split("=", 1)
        rows.append([key.strip(), value.strip(), comment, "可调模型参数", "参与首追时间、概率或说明计算"])
        comment = ""
    return rows


def add_sheet(wb, title, headers, rows):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for column_cells in ws.columns:
        max_len = 0
        column = column_cells[0].column
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[get_column_letter(column)].width = min(max(max_len + 2, 10), 48)
    ws.freeze_panes = "A2"


def main():
    wb = Workbook()
    wb.remove(wb.active)

    add_sheet(wb, "求生者",
              ["编号", "名称", "破译原始", "破译百分", "牵制原始", "牵制百分", "辅助原始", "辅助百分", "救援原始", "救援百分", "首抓原始", "首抓百分", "标签", "建模说明"],
              parse_role_file("survivor_data.h"))
    add_sheet(wb, "监管者",
              ["编号", "名称", "追击原始", "追击百分", "守椅原始", "守椅百分", "控场原始", "控场百分", "信息原始", "信息百分", "强度原始", "强度百分", "标签", "建模说明"],
              parse_role_file("hunter_data.h"))
    add_sheet(wb, "辅助技能",
              ["编号", "名称", "开局CD秒", "首刀修正秒", "倒地修正秒", "标签", "技能作用", "怎么作用于计算"],
              parse_auxiliary_traits())

    map_rows = parse_maps()
    add_sheet(wb, "地图区域", ["地图编号", "地图", "区域编号", "区域", "X", "Y", "宽", "高", "牵制分", "开阔分", "电机价值", "区域说明"], map_rows)
    add_sheet(wb, "地图统计", ["地图编号", "地图", "类型", "区域数量", "方位口径"], build_map_stats(map_rows))
    add_sheet(wb, "几何地图", ["地图编号", "地图键", "图源文件", "网格宽", "网格高", "寻路宽度m", "寻路高度m", "墙格数", "窗数量", "板数量", "说明"], build_geometry_rows())
    add_sheet(wb, "参数文件", ["参数名", "当前值", "注释", "参数是什么", "怎么作用于计算"], parse_model_params())
    add_sheet(wb, "资料来源", ["来源", "链接", "用途"], [
        ["第五人格 BWIKI：辅助特质", "https://wiki.biligame.com/dwrg/辅助特质", "辅助技能、开局冷却和效果口径"],
        ["第五人格 BWIKI：角色数据", "https://wiki.biligame.com/dwrg/角色数据", "角色和移动速度口径"],
        ["用户提供地图原图", "data/map/*.jpg", "地图显示、几何识别和点选复核"],
        ["参数文件", "config/model_params.txt", "首追模型可调权重"],
    ])

    output = root / "data" / "identityv_scores.xlsx"
    output.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(output)
        print(output)
    except PermissionError:
        fallback = root / "data" / "identityv_scores_updated.xlsx"
        wb.save(fallback)
        print(fallback)


if __name__ == "__main__":
    main()
